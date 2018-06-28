import sys, csv, django, os, codecs, re, fnmatch

#sys.path.append('/home/galm/software/django/tmv/BasicBrowser')
sys.path.append('/home/hilj/github/tmv/BasicBrowser')

# sys.path.append('/home/max/Desktop/django/BasicBrowser/')
import db as db
from tmv_app.models import *
from scoping.models import *
from scoping.views import *

from utils.utils import *
from utils.text import *
import xml.etree.ElementTree as ET
import pandas as pd
import short_url
from django.core.files import File
import openpyxl

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "BasicBrowser.settings")
django.setup()

book = openpyxl.load_workbook('/home/galm/projects/NETs-IAMs/parsed_pars_and_manualupdate.xslx.xlsx')

sheet = book['missing_pars']

rows = sheet.max_row
cols = sheet.max_column
headers = dict((i, sheet.cell(row=1, column=i).value) for i in range(1, cols))
def item(i, j):
    return (sheet.cell(row=1, column=j).value, sheet.cell(row=i, column=j).value)

data = (dict(item(i, j) for j in range(1, cols + 1)) for i in range(2, rows + 1))
ldoc = None
for row in data:
    row = row
    try:
        doc = Doc.objects.get(title=row['doc__title'])
    except:
        print(row['doc__title'])
        continue
    if doc !=ldoc:
        n = doc.docpar_set.count()
    else:
        n+=1
        ldoc=doc
    dp = DocPar(
        doc=doc,
        text=row['text'],
        n=n
    )
    dp.save()
